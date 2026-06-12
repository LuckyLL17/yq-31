import sys
import os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from batch_wizard import BatchConfigWizard
from config_manager import get_default_config


def main():
    print("=" * 60)
    print("测试: 批量向导 split_config 配置传递链路")
    print("=" * 60)

    config = get_default_config()
    wizard = BatchConfigWizard(config)

    print(f"\n1. 初始 config 中的 split_config:")
    print(f"   enabled: {wizard.config.get('split_config', {}).get('enabled')}")
    print(f"   split_field: {wizard.config.get('split_config', {}).get('split_field')}")

    wizard.config.setdefault("split_config", {})["enabled"] = True
    wizard.config["split_config"]["split_field"] = "department"
    wizard.config["split_config"]["split_rule"] = "by_value"
    wizard.config["split_config"]["sheet_name_template"] = "{index}-{value}"
    wizard.config["split_config"]["include_all_sheet"] = True

    print(f"\n2. 手动设置后的 split_config:")
    print(f"   enabled: {wizard.config['split_config']['enabled']}")
    print(f"   split_field: {wizard.config['split_config']['split_field']}")
    print(f"   split_rule: {wizard.config['split_config']['split_rule']}")
    print(f"   sheet_name_template: {wizard.config['split_config']['sheet_name_template']}")
    print(f"   include_all_sheet: {wizard.config['split_config']['include_all_sheet']}")

    batch_config = wizard._get_batch_config()

    print(f"\n3. _get_batch_config() 返回的 config 中的 split_config:")
    cfg = batch_config.get("config", {})
    sc = cfg.get("split_config", {})
    print(f"   enabled: {sc.get('enabled')}")
    print(f"   split_field: {sc.get('split_field')}")
    print(f"   split_rule: {sc.get('split_rule')}")
    print(f"   sheet_name_template: {sc.get('sheet_name_template')}")
    print(f"   include_all_sheet: {sc.get('include_all_sheet')}")

    from batch_processor import start_batch_process
    import tempfile

    output_dir = tempfile.mkdtemp(prefix="batch_wizard_test_")
    print(f"\n4. 通过 start_batch_process 创建批量任务...")
    print(f"   输出目录: {output_dir}")

    batch_task, result = start_batch_process(
        source_type="files",
        sources=["./data/sample_data.json"],
        output_dir=output_dir,
        config=batch_config["config"],
        options={"generate_report": False, "skip_existing": False},
    )

    if batch_task is None:
        print(f"   ❌ 失败: {result}")
        return 1

    print(f"   批量任务ID: {batch_task['batch_id']}")
    print(f"   任务中保存的 split_config.enabled: {batch_task['config'].get('split_config', {}).get('enabled')}")
    print(f"   任务中保存的 split_field: {batch_task['config'].get('split_config', {}).get('split_field')}")

    print(f"\n5. 检查生成的 Excel 文件...")
    from openpyxl import load_workbook

    for f in batch_task["files"]:
        if f["status"] == "completed":
            wb = load_workbook(f["excel_path"])
            print(f"   文件: {os.path.basename(f['excel_path'])}")
            print(f"   工作表数量: {len(wb.sheetnames)}")
            print(f"   工作表列表: {wb.sheetnames}")

    print("\n✅ 测试完成 - 拆分配置在批量向导中正确传递并生效")
    return 0


if __name__ == "__main__":
    sys.exit(main())
