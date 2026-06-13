import json
import os
import sys
import copy
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.comments import Comment
from datetime import datetime

from config_manager import (
    load_config,
    save_config,
    get_default_config,
    validate_config,
)
from data_validator import (
    ValidationRule,
    validate_data,
    apply_validation_to_export,
    rules_from_config,
    ON_FAIL_MARK,
    ON_FAIL_ABORT,
    MARK_COLOR,
    MARK_BG_COLOR,
    ON_FAIL_LABELS,
)


def load_json(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"JSON文件不存在: {file_path}")

    with open(file_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if isinstance(data, dict):
        if "data" in data and isinstance(data["data"], list):
            data = data["data"]
        elif "items" in data and isinstance(data["items"], list):
            data = data["items"]
        elif "records" in data and isinstance(data["records"], list):
            data = data["records"]
        else:
            data = [data]
    elif not isinstance(data, list):
        data = [data]

    return data


def flatten_dict(d, parent_key="", sep="."):
    items = []
    if isinstance(d, dict):
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(flatten_dict(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                items.append((new_key, json.dumps(v, ensure_ascii=False)))
            else:
                items.append((new_key, v))
    else:
        items.append((parent_key, d))
    return dict(items)


def auto_detect_headers(data):
    headers = []
    all_keys = set()
    for item in data:
        if isinstance(item, dict):
            flat = flatten_dict(item)
            all_keys.update(flat.keys())

    for key in sorted(all_keys):
        headers.append({
            "key": key,
            "label": key.replace(".", " / ").replace("_", " ").title(),
            "width": max(15, len(key) * 2 + 5),
        })
    return headers


def merge_headers(config_headers, auto_headers, config):
    if not config_headers:
        return auto_headers

    config_keys = {h["key"] for h in config_headers}
    extra_headers = [h for h in auto_headers if h["key"] not in config_keys]

    if config.get("auto_detect_headers", True) and extra_headers:
        return config_headers + extra_headers
    return config_headers


def _get_alignment(align_str):
    align_map = {
        "left": "left",
        "center": "center",
        "right": "right",
    }
    return align_map.get(align_str, "center")


def _create_border(style_config):
    border_style = style_config.get("border_style", "thin")
    border_color = style_config.get("border_color", "000000").replace("#", "")

    if border_style is None:
        return None

    side = Side(style=border_style, color=border_color)
    return Border(left=side, right=side, top=side, bottom=side)


def apply_header_style(ws, header_cells, config):
    header_style = config.get("header_style", {})

    font_name = header_style.get("font_name", "Microsoft YaHei")
    font_bold = header_style.get("font_bold", True)
    font_color = header_style.get("font_color", "FFFFFF").replace("#", "")
    font_size = header_style.get("font_size", 12)
    bg_color = header_style.get("bg_color", "4472C4").replace("#", "")
    alignment = _get_alignment(header_style.get("alignment", "center"))
    vertical_alignment = header_style.get("vertical_alignment", "center")

    font = Font(name=font_name, bold=font_bold, color=font_color, size=font_size)
    fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    alignment = Alignment(horizontal=alignment, vertical=vertical_alignment)
    border = _create_border(header_style)

    for cell in header_cells:
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment
        if border:
            cell.border = border


def apply_data_style(ws, headers, data_rows_count, config):
    data_style = config.get("data_style", {})
    font_name = data_style.get("font_name", "Microsoft YaHei")
    font_bold = data_style.get("font_bold", False)
    font_color = data_style.get("font_color", "000000").replace("#", "")
    font_size = data_style.get("font_size", 11)
    wrap_text = data_style.get("wrap_text", True)
    alt_row_color = data_style.get("alt_row_color", "F2F2F2").replace("#", "")
    alignment = _get_alignment(data_style.get("alignment", "left"))
    vertical_alignment = data_style.get("vertical_alignment", "center")

    font = Font(name=font_name, bold=font_bold, color=font_color, size=font_size)
    alignment = Alignment(horizontal=alignment, vertical=vertical_alignment, wrap_text=wrap_text)
    border = _create_border(data_style)
    alt_fill = PatternFill(start_color=alt_row_color, end_color=alt_row_color, fill_type="solid")

    for row_idx in range(2, 2 + data_rows_count):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font
            cell.alignment = alignment
            if border:
                cell.border = border
            if config.get("style_alt_rows", True) and row_idx % 2 == 0:
                cell.fill = alt_fill


def set_column_widths(ws, headers):
    for idx, header in enumerate(headers, start=1):
        width = header.get("width", 15)
        col_letter = chr(64 + idx) if idx <= 26 else _get_column_letter(idx)
        ws.column_dimensions[col_letter].width = width


def _get_column_letter(col_idx):
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def extract_value(item, key):
    flat = flatten_dict(item)
    if key in flat:
        value = flat[key]
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False)
        return value

    if "." in key:
        parts = key.split(".")
        current = item
        for part in parts:
            if isinstance(current, dict) and part in current:
                current = current[part]
            else:
                return ""
        if isinstance(current, (dict, list)):
            return json.dumps(current, ensure_ascii=False)
        return current

    return ""


def _sanitize_sheet_name(name, max_length=31):
    invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
    result = str(name)
    for ch in invalid_chars:
        result = result.replace(ch, '_')
    result = result.strip()
    if not result:
        result = "Sheet"
    if len(result) > max_length:
        result = result[:max_length]
    return result


def _render_sheet_name(template, value, index, total_count, empty_label="未分类"):
    display_value = value if value is not None and value != "" else empty_label
    try:
        return template.format(
            value=display_value,
            index=index,
            count=total_count,
            num=index,
        )
    except (KeyError, IndexError):
        return str(display_value)


def _match_custom_rule(value, rule):
    if "values" in rule:
        rule_values = rule["values"]
        if isinstance(rule_values, list):
            return value in rule_values
        return value == rule_values

    if "condition" in rule:
        try:
            return eval(rule["condition"], {"__builtins__": {}}, {"value": value})
        except Exception:
            return False

    if "min" in rule or "max" in rule:
        try:
            num_val = float(value) if value is not None and value != "" else 0
            min_val = rule.get("min")
            max_val = rule.get("max")
            include_min = rule.get("include_min", True)
            include_max = rule.get("include_max", False)

            min_ok = True
            if min_val is not None:
                if include_min:
                    min_ok = num_val >= min_val
                else:
                    min_ok = num_val > min_val

            max_ok = True
            if max_val is not None:
                if include_max:
                    max_ok = num_val <= max_val
                else:
                    max_ok = num_val < max_val

            return min_ok and max_ok
        except (ValueError, TypeError):
            return False

    return False


def _match_range_group(value, group):
    try:
        num_val = float(value) if value is not None and value != "" else None
        if num_val is None:
            return False

        min_val = group.get("min")
        max_val = group.get("max")
        include_min = group.get("include_min", True)
        include_max = group.get("include_max", False)

        min_ok = True
        if min_val is not None:
            if include_min:
                min_ok = num_val >= min_val
            else:
                min_ok = num_val > min_val

        max_ok = True
        if max_val is not None:
            if include_max:
                max_ok = num_val <= max_val
            else:
                max_ok = num_val < max_val

        return min_ok and max_ok
    except (ValueError, TypeError):
        return False


def split_data_by_field(data, split_field, split_config):
    split_rule = split_config.get("split_rule", "by_value")
    empty_label = split_config.get("empty_value_label", "未分类")

    groups = []
    group_order = []
    group_to_index = {}

    def _get_or_create_group(name):
        if name not in group_to_index:
            idx = len(groups)
            groups.append({"name": name, "data": [], "original_indices": []})
            group_order.append(name)
            group_to_index[name] = idx
        return group_to_index[name]

    for orig_idx, item in enumerate(data):
        if not isinstance(item, dict):
            continue
        value = extract_value(item, split_field)

        matched = False

        if split_rule == "by_value":
            group_name = value if value is not None and value != "" else empty_label
            gidx = _get_or_create_group(group_name)
            groups[gidx]["data"].append(item)
            groups[gidx]["original_indices"].append(orig_idx)
            matched = True

        elif split_rule == "by_range":
            range_groups = split_config.get("range_groups", [])
            for rg in range_groups:
                if _match_range_group(value, rg):
                    gidx = _get_or_create_group(rg["name"])
                    groups[gidx]["data"].append(item)
                    groups[gidx]["original_indices"].append(orig_idx)
                    matched = True
                    break
            if not matched:
                gidx = _get_or_create_group(empty_label)
                groups[gidx]["data"].append(item)
                groups[gidx]["original_indices"].append(orig_idx)

        elif split_rule == "by_custom":
            custom_rules = split_config.get("custom_rules", [])
            for cr in custom_rules:
                if _match_custom_rule(value, cr):
                    gidx = _get_or_create_group(cr["name"])
                    groups[gidx]["data"].append(item)
                    groups[gidx]["original_indices"].append(orig_idx)
                    matched = True
                    break
            if not matched:
                fallback_name = split_config.get("fallback_group_name", empty_label)
                gidx = _get_or_create_group(fallback_name)
                groups[gidx]["data"].append(item)
                groups[gidx]["original_indices"].append(orig_idx)

    return groups


def _write_sheet_data(ws, data, headers, config, original_indices=None, validation_result=None):
    header_labels = [h["label"] for h in headers]
    ws.append(header_labels)

    for item in data:
        if not isinstance(item, dict):
            continue
        row = [extract_value(item, h["key"]) for h in headers]
        ws.append(row)

    set_column_widths(ws, headers)

    if config.get("style_header", True):
        header_cells = ws[1]
        apply_header_style(ws, header_cells, config)

    apply_data_style(ws, headers, len(data), config)

    if validation_result and validation_result.marked_rows and original_indices is not None:
        _apply_validation_marks(ws, validation_result, headers, original_indices)

    ws.freeze_panes = "A2"


def export_to_excel_with_split(data, headers, config, validation_result=None, original_indices=None):
    output_path = config["excel_output_path"]
    sheet_name = config.get("sheet_name", "数据导出")
    split_config = config.get("split_config", {})
    split_field = split_config["split_field"]

    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    groups = split_data_by_field(data, split_field, split_config)
    total_groups = len(groups)

    print(f"\n按字段 '{split_field}' 拆分为 {total_groups} 个分组:")
    for gi, g in enumerate(groups):
        print(f"  [{gi + 1}] {g['name']}: {len(g['data'])} 条数据")
    print()

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    sheet_name_template = split_config.get("sheet_name_template", "{value}")
    max_sheet_len = split_config.get("max_sheet_name_length", 31)
    include_all = split_config.get("include_all_sheet", True)
    all_sheet_name = split_config.get("all_sheet_name", "全部数据")

    if include_all:
        ws_all = wb.create_sheet(_sanitize_sheet_name(all_sheet_name, max_sheet_len))
        _write_sheet_data(
            ws_all,
            data,
            headers,
            config,
            original_indices=original_indices,
            validation_result=validation_result,
        )

    used_names = set()
    for gi, group in enumerate(groups, start=1):
        raw_name = _render_sheet_name(
            sheet_name_template,
            group["name"],
            gi,
            total_groups,
            split_config.get("empty_value_label", "未分类"),
        )
        safe_name = _sanitize_sheet_name(raw_name, max_sheet_len)

        base_name = safe_name
        suffix = 1
        while safe_name in used_names:
            suffix += 1
            extra = f"_{suffix}"
            if len(base_name) + len(extra) > max_sheet_len:
                safe_name = base_name[: max_sheet_len - len(extra)] + extra
            else:
                safe_name = base_name + extra

        used_names.add(safe_name)

        ws = wb.create_sheet(safe_name)

        group_orig_indices = None
        if original_indices is not None:
            group_orig_indices = [original_indices[i] for i in group["original_indices"]]

        _write_sheet_data(
            ws,
            group["data"],
            headers,
            config,
            original_indices=group_orig_indices,
            validation_result=validation_result,
        )

    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet(_sanitize_sheet_name(sheet_name, max_sheet_len))
        _write_sheet_data(ws, data, headers, config)

    add_pivot_table_to_workbook(wb, data, headers, config)

    wb.save(output_path)
    print(f"Excel文件已成功导出: {os.path.abspath(output_path)}")
    print(f"共 {len(wb.sheetnames)} 个工作表，{len(data)} 条数据，{len(headers)} 个字段")
    if validation_result:
        skipped_count = len(validation_result.skipped_rows)
        marked_count = len(validation_result.marked_rows)
        if skipped_count > 0:
            print(f"跳过 {skipped_count} 条", end="")
        if marked_count > 0:
            print(f"，标记 {marked_count} 条", end="")
        if skipped_count > 0 or marked_count > 0:
            print()
    return output_path


def export_to_excel(data, headers, config):
    output_path = config["excel_output_path"]
    sheet_name = config.get("sheet_name", "数据导出")

    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    rules = rules_from_config(config)
    validation_result = None
    original_data = data
    original_indices = list(range(len(data)))

    if rules:
        print(f"\n正在执行数据校验（{len(rules)} 条规则）...")
        validation_result = validate_data(data, rules)
        print(validation_result.summary())

        if validation_result.aborted:
            print(f"\n❌ 校验中止，导出已取消: {validation_result.abort_reason}")
            _print_validation_errors(validation_result, data, max_display=10)
            return None

        if validation_result.skipped_rows:
            print(f"  将跳过 {len(validation_result.skipped_rows)} 行校验不通过的数据")

        valid_data, removed = apply_validation_to_export(data, validation_result, headers)
        if valid_data is None:
            print("\n❌ 校验中止，导出已取消")
            return None

        new_indices = []
        for idx in range(len(data)):
            if idx not in validation_result.skipped_rows:
                new_indices.append(idx)
        original_indices = new_indices
        data = valid_data
        print(f"  校验后有效数据: {len(data)} 条\n")

    split_config = config.get("split_config", {})
    if split_config.get("enabled") and split_config.get("split_field"):
        return export_to_excel_with_split(
            data=data,
            headers=headers,
            config=config,
            validation_result=validation_result,
            original_indices=original_indices,
        )

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_labels = [h["label"] for h in headers]
    ws.append(header_labels)

    for item in data:
        if not isinstance(item, dict):
            continue
        row = [extract_value(item, h["key"]) for h in headers]
        ws.append(row)

    set_column_widths(ws, headers)

    if config.get("style_header", True):
        header_cells = ws[1]
        apply_header_style(ws, header_cells, config)

    apply_data_style(ws, headers, len(data), config)

    if validation_result and validation_result.marked_rows:
        _apply_validation_marks(ws, validation_result, headers, original_indices)

    ws.freeze_panes = "A2"

    add_pivot_table_to_workbook(wb, data, headers, config)

    wb.save(output_path)
    print(f"Excel文件已成功导出: {os.path.abspath(output_path)}")
    if validation_result:
        skipped_count = len(validation_result.skipped_rows)
        marked_count = len(validation_result.marked_rows)
        print(f"共导出 {len(data)} 条数据，{len(headers)} 个字段", end="")
        if skipped_count > 0:
            print(f"，跳过 {skipped_count} 条", end="")
        if marked_count > 0:
            print(f"，标记 {marked_count} 条", end="")
        print()
    else:
        print(f"共导出 {len(data)} 条数据，{len(headers)} 个字段")
    return output_path


def _apply_validation_marks(ws, validation_result, headers, original_indices):
    mark_font = Font(color=MARK_COLOR, bold=True)
    mark_fill = PatternFill(start_color=MARK_BG_COLOR, end_color=MARK_BG_COLOR, fill_type="solid")

    header_key_to_col = {}
    for col_idx, header in enumerate(headers, start=1):
        header_key_to_col[header["key"]] = col_idx

    marked_in_export = set()
    for orig_idx in validation_result.marked_rows:
        if orig_idx in original_indices:
            export_row = original_indices.index(orig_idx) + 2
            marked_in_export.add(export_row)

            row_errors = validation_result.get_field_errors_for_row(orig_idx)
            error_msgs = []
            for field_key, errors in row_errors.items():
                for e in errors:
                    error_msgs.append(e.rule.message)

            comment_text = "; ".join(error_msgs)
            if len(comment_text) > 200:
                comment_text = comment_text[:200] + "..."

            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=export_row, column=col_idx)
                cell.font = mark_font
                cell.fill = mark_fill

            first_cell = ws.cell(row=export_row, column=1)
            first_cell.comment = Comment(comment_text, "数据校验")

            for field_key, errors in row_errors.items():
                if field_key in header_key_to_col:
                    col_idx = header_key_to_col[field_key]
                    cell = ws.cell(row=export_row, column=col_idx)
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")


def _print_validation_errors(validation_result, data, max_display=10):
    errors = validation_result.errors[:max_display]
    if not errors:
        return
    print("\n校验错误详情:")
    for e in errors:
        val_str = str(e.value)[:30] if e.value is not None else "(空)"
        on_fail_str = ON_FAIL_LABELS.get(e.rule.on_fail, "").split("（")[0]
        print(f"  行 {e.row_index + 1}: {e.rule.message} | 值: {val_str} | 处理: {on_fail_str}")
    if len(validation_result.errors) > max_display:
        print(f"  ... 还有 {len(validation_result.errors) - max_display} 个问题未显示")


def parse_args():
    parser = argparse.ArgumentParser(description="JSON 数据导出工具（支持 Excel/CSV/TSV/HTML/Markdown/JSON/PDF）")
    parser.add_argument(
        "-w", "--wizard",
        action="store_true",
        help="使用交互式配置向导",
    )
    parser.add_argument(
        "-p", "--preview",
        action="store_true",
        help="进入数据预览与筛选模式",
    )
    parser.add_argument(
        "-c", "--config",
        type=str,
        default=None,
        help="指定配置文件路径（默认: ./config.json）",
    )
    parser.add_argument(
        "-i", "--input",
        type=str,
        default=None,
        help="指定输入 JSON 文件路径",
    )
    parser.add_argument(
        "-o", "--output",
        type=str,
        default=None,
        help="指定输出文件路径",
    )
    parser.add_argument(
        "-f", "--format",
        type=str,
        default=None,
        choices=["excel", "csv", "tsv", "html", "markdown", "json", "pdf"],
        help="指定导出格式 (excel/csv/tsv/html/markdown/json/pdf)",
    )
    parser.add_argument(
        "--list-formats",
        action="store_true",
        help="列出所有支持的导出格式",
    )
    parser.add_argument(
        "--save-config",
        type=str,
        default=None,
        help="将当前配置保存到指定文件",
    )
    parser.add_argument(
        "--list-templates",
        action="store_true",
        help="列出所有可用的样式模板",
    )
    parser.add_argument(
        "--apply-template",
        type=str,
        default=None,
        help="应用指定的样式模板",
    )
    parser.add_argument(
        "--template-manager",
        action="store_true",
        help="启动样式模板管理器",
    )
    parser.add_argument(
        "--save-as-template",
        type=str,
        default=None,
        help="将当前样式保存为模板（指定模板ID）",
    )
    parser.add_argument(
        "--template-name",
        type=str,
        default=None,
        help="保存模板时的名称（与 --save-as-template 配合使用）",
    )
    parser.add_argument(
        "--template-desc",
        type=str,
        default="",
        help="保存模板时的描述（与 --save-as-template 配合使用）",
    )
    parser.add_argument(
        "--batch",
        action="store_true",
        help="使用批量处理模式（交互式向导）",
    )
    parser.add_argument(
        "--batch-dir",
        type=str,
        default=None,
        help="批量处理：指定包含JSON文件的目录",
    )
    parser.add_argument(
        "--batch-files",
        type=str,
        default=None,
        help="批量处理：指定多个JSON文件，用逗号分隔",
    )
    parser.add_argument(
        "--batch-output",
        type=str,
        default=None,
        help="批量处理：指定输出目录",
    )
    parser.add_argument(
        "--batch-resume",
        type=str,
        default=None,
        help="批量处理：恢复指定ID的未完成任务",
    )
    parser.add_argument(
        "--batch-list",
        action="store_true",
        help="批量处理：列出所有历史任务",
    )
    parser.add_argument(
        "--batch-report",
        type=str,
        default=None,
        help="批量处理：生成指定任务ID的处理报告",
    )
    parser.add_argument(
        "--batch-report-format",
        type=str,
        default="txt",
        choices=["txt", "json", "both"],
        help="批量处理：报告格式 (txt/json/both)，默认txt",
    )
    parser.add_argument(
        "--batch-skip-existing",
        action="store_true",
        help="批量处理：跳过已存在的Excel文件",
    )
    parser.add_argument(
        "--validate",
        action="store_true",
        help="启动交互式数据校验规则配置",
    )
    parser.add_argument(
        "--validate-only",
        action="store_true",
        help="仅执行数据校验，不导出Excel",
    )
    return parser.parse_args()


def apply_cli_overrides(config, args):
    from multi_exporter import get_format_extension

    if args.input:
        config["json_file_path"] = args.input
    if args.format:
        config["export_format"] = args.format
    if args.output:
        fmt = config.get("export_format", "excel")
        if fmt == "excel":
            config["excel_output_path"] = args.output
        else:
            config[f"{fmt}_output_path"] = args.output
    return config


def run_with_config(config, data=None):
    from multi_exporter import export_data, EXPORT_FORMATS

    export_fmt = config.get("export_format", "excel")
    fmt_info = EXPORT_FORMATS.get(export_fmt, {})
    output_key = f"{export_fmt}_output_path" if export_fmt != "excel" else "excel_output_path"

    print("=" * 50)
    print(f"JSON 数据导出工具 - {fmt_info.get('label', export_fmt.upper())}")
    print("=" * 50)

    json_path = config["json_file_path"]
    output_path = config.get(output_key, "")

    print(f"JSON文件路径: {json_path}")
    print(f"输出格式: {fmt_info.get('label', export_fmt.upper())}")
    print(f"输出路径: {output_path}")
    print()

    try:
        if data is None:
            data = load_json(json_path)
            print(f"已加载 {len(data)} 条数据")
        else:
            print(f"使用筛选后的数据，共 {len(data)} 条")

        auto_headers = auto_detect_headers(data)
        print(f"自动检测到 {len(auto_headers)} 个字段")

        headers = merge_headers(config.get("default_headers", []), auto_headers, config)
        print(f"最终使用 {len(headers)} 个字段")
        print()

        export_data(
            data=data,
            headers=headers,
            config=config,
            fmt=export_fmt,
        )

        if export_fmt == "excel":
            prompt_save_as_template(config)

    except FileNotFoundError as e:
        print(f"错误: {e}")
        print("请检查JSON文件路径配置是否正确")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"JSON解析错误: {e}")
        print("请检查JSON文件格式是否正确")
        sys.exit(1)
    except Exception as e:
        print(f"发生错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


def prompt_save_as_template(config):
    from prompts import prompt_confirm, prompt_input
    from style_template_manager import create_template_from_config

    if not prompt_confirm("\n是否将当前样式保存为自定义模板？", default=False):
        return

    template_id = prompt_input("请输入模板ID (英文标识，如 'my_style')", required=True)
    name = prompt_input("请输入模板名称", required=True)
    description = prompt_input("请输入模板描述", required=False, default="")

    success, msg = create_template_from_config(
        template_id, name, description, config, overwrite=False
    )

    if not success and "已存在" in msg:
        if prompt_confirm(f"模板 '{template_id}' 已存在，是否覆盖？", default=False):
            success, msg = create_template_from_config(
                template_id, name, description, config, overwrite=True
            )

    print(f"\n{msg}")


def handle_template_operations(args, config):
    from style_template_manager import (
        list_templates,
        apply_template_to_config,
        create_template_from_config,
        get_border_style_name,
    )

    if args.list_templates:
        templates = list_templates()
        print("=" * 70)
        print("可用样式模板")
        print("=" * 70)
        for template_id, template in templates.items():
            is_builtin = template_id in ("default", "professional", "fresh", "warm", "elegant", "minimal")
            builtin_tag = " [内置]" if is_builtin else " [自定义]"
            header_bg = template.get("header_style", {}).get("bg_color", "#FFFFFF")
            alt_color = template.get("data_style", {}).get("alt_row_color", "#FFFFFF")
            border_style = template.get("header_style", {}).get("border_style", "thin")
            border_name = get_border_style_name(border_style) if border_style else "无"
            print(f"\n{template_id}{builtin_tag}")
            print(f"  名称: {template.get('name', '未命名')}")
            print(f"  描述: {template.get('description', '无描述')}")
            print(f"  表头背景: {header_bg} | 隔行色: {alt_color} | 边框: {border_name}")
        return True

    if args.template_manager:
        from style_template_wizard import run_style_template_manager
        run_style_template_manager(config)
        return True

    if args.apply_template:
        success, msg = apply_template_to_config(config, args.apply_template)
        print(msg)
        if not success:
            sys.exit(1)
        return False

    if args.save_as_template:
        template_id = args.save_as_template
        name = args.template_name or template_id
        description = args.template_desc
        success, msg = create_template_from_config(
            template_id, name, description, config, overwrite=False
        )
        print(msg)
        if not success:
            sys.exit(1)
        return True

    return False


def _handle_batch_operations(args):
    if args.batch:
        from batch_wizard import run_batch_wizard
        config = load_config(args.config) if args.config else get_default_config()
        run_batch_wizard(config)
        return True

    if args.batch_list:
        from batch_wizard import show_task_manager
        show_task_manager()
        return True

    if args.batch_resume:
        from batch_processor import resume_batch_process
        from report_generator import generate_report, generate_json_report, print_summary

        batch_id = args.batch_resume
        batch_task, result = resume_batch_process(batch_id)

        if batch_task is None:
            print(f"错误: {result}")
            sys.exit(1)

        fmt = args.batch_report_format
        if fmt in ("txt", "both"):
            report_path = generate_report(batch_task)
            print(f"\n📄 文本报告已生成: {os.path.abspath(report_path)}")
        if fmt in ("json", "both"):
            report_path = generate_json_report(batch_task)
            print(f"📄 JSON报告已生成: {os.path.abspath(report_path)}")

        print_summary(batch_task)
        return True

    if args.batch_report:
        from task_manager import load_batch_task
        from report_generator import generate_report, generate_json_report, print_summary

        batch_id = args.batch_report
        batch_task = load_batch_task(batch_id)
        if not batch_task:
            print(f"错误: 任务 {batch_id} 不存在")
            sys.exit(1)

        fmt = args.batch_report_format
        if fmt in ("txt", "both"):
            report_path = generate_report(batch_task)
            print(f"📄 文本报告已生成: {os.path.abspath(report_path)}")
        if fmt in ("json", "both"):
            report_path = generate_json_report(batch_task)
            print(f"📄 JSON报告已生成: {os.path.abspath(report_path)}")

        print_summary(batch_task)
        return True

    if args.batch_dir or args.batch_files:
        from batch_processor import start_batch_process
        from report_generator import generate_report, generate_json_report, print_summary

        config = load_config(args.config)
        config = apply_cli_overrides(config, args)

        output_dir = args.batch_output or "./output/batch"

        if args.batch_dir:
            source_type = "directory"
            sources = [args.batch_dir]
        else:
            source_type = "files"
            sources = [f.strip() for f in args.batch_files.split(",") if f.strip()]

        batch_task, result = start_batch_process(
            source_type=source_type,
            sources=sources,
            output_dir=output_dir,
            config=config,
            options={
                "generate_report": True,
                "report_format": args.batch_report_format,
                "skip_existing": args.batch_skip_existing,
            },
        )

        if batch_task is None:
            print(f"错误: {result}")
            sys.exit(1)

        fmt = args.batch_report_format
        if fmt in ("txt", "both"):
            report_path = generate_report(batch_task)
            print(f"\n📄 文本报告已生成: {os.path.abspath(report_path)}")
        if fmt in ("json", "both"):
            report_path = generate_json_report(batch_task)
            print(f"📄 JSON报告已生成: {os.path.abspath(report_path)}")

        print_summary(batch_task)
        return True

    return False


def main():
    args = parse_args()

    if args.list_formats:
        from multi_exporter import EXPORT_FORMATS
        print("=" * 70)
        print("支持的导出格式")
        print("=" * 70)
        for fmt_id, fmt_info in EXPORT_FORMATS.items():
            print(f"\n  {fmt_id:10s} - {fmt_info['label']}")
            print(f"              {fmt_info['description']}")
        print("\n使用 -f/--format 参数指定格式，例如: python json_to_excel.py -f csv")
        return

    if _handle_batch_operations(args):
        return

    if args.validate:
        from validation_wizard import run_validation_wizard
        config = load_config(args.config) if args.config else get_default_config()
        try:
            data = load_json(config["json_file_path"])
            auto_headers = auto_detect_headers(data)
            config = run_validation_wizard(config, available_fields=auto_headers)
        except Exception:
            config = run_validation_wizard(config)
        if args.save_config:
            save_config(config, args.save_config)
            print(f"\n配置已保存到: {os.path.abspath(args.save_config)}")
        return

    if args.validate_only:
        config = load_config(args.config) if args.config else get_default_config()
        config = apply_cli_overrides(config, args)
        rules = rules_from_config(config)
        if not rules:
            print("未配置任何校验规则，请先使用 --validate 配置规则")
            return
        try:
            data = load_json(config["json_file_path"])
            print(f"已加载 {len(data)} 条数据，正在执行校验...")
            result = validate_data(data, rules)
            print(f"\n{result.summary()}")
            _print_validation_errors(result, data)
        except Exception as e:
            print(f"校验执行失败: {e}")
        return

    if args.wizard:
        from wizard import run_wizard
        config = load_config(args.config) if args.config else get_default_config()
        config = run_wizard(config)

        errors = validate_config(config)
        if errors:
            print("\n❌ 配置验证失败:")
            for e in errors:
                print(f"  • {e}")
            sys.exit(1)

        if args.save_config:
            save_config(config, args.save_config)
            print(f"\n配置已保存到: {os.path.abspath(args.save_config)}")

        if prompt_confirm_execute():
            filtered_data = config.pop("_filtered_data", None)
            run_with_config(config, data=filtered_data)
        return

    config = load_config(args.config)
    config = apply_cli_overrides(config, args)

    template_exit = handle_template_operations(args, config)
    if template_exit and not args.apply_template:
        return

    if args.apply_template:
        if args.save_config:
            save_config(config, args.save_config)
            print(f"配置已保存到: {os.path.abspath(args.save_config)}")
        if prompt_confirm_execute():
            run_with_config(config)
        return

    if args.preview:
        from data_previewer import start_preview_mode
        from multi_exporter import export_data
        result = start_preview_mode(config)
        if result is not None and prompt_confirm_execute_after_preview(config):
            export_config = copy.deepcopy(config)
            auto_headers = auto_detect_headers(result)
            headers = merge_headers(config.get("default_headers", []), auto_headers, config)
            export_fmt = config.get("export_format", "excel")
            export_data(data=result, headers=headers, config=export_config, fmt=export_fmt)
            if export_fmt == "excel":
                prompt_save_as_template(export_config)
        return

    errors = validate_config(config)
    if errors:
        print("❌ 配置验证失败:")
        for e in errors:
            print(f"  • {e}")
        sys.exit(1)

    if args.save_config:
        save_config(config, args.save_config)
        print(f"配置已保存到: {os.path.abspath(args.save_config)}")

    if prompt_confirm_preview():
        from data_previewer import start_preview_mode
        from multi_exporter import export_data
        result = start_preview_mode(config)
        if result is not None and prompt_confirm_execute_after_preview(config):
            export_config = copy.deepcopy(config)
            auto_headers = auto_detect_headers(result)
            headers = merge_headers(config.get("default_headers", []), auto_headers, config)
            export_fmt = config.get("export_format", "excel")
            export_data(data=result, headers=headers, config=export_config, fmt=export_fmt)
            if export_fmt == "excel":
                prompt_save_as_template(export_config)
        return

    run_with_config(config)


def prompt_confirm_preview():
    while True:
        user_input = input("\n是否先预览并筛选数据？ (Y/n): ").strip().lower()
        if not user_input or user_input in ("y", "yes"):
            return True
        if user_input in ("n", "no"):
            return False
        print("  ⚠️  请输入 y 或 n")


def prompt_confirm_execute():
    while True:
        user_input = input("\n配置完成！是否立即执行导出？ (Y/n): ").strip().lower()
        if not user_input or user_input in ("y", "yes"):
            return True
        if user_input in ("n", "no"):
            print("配置已保存，您可以稍后运行 `python json_to_excel.py` 执行导出。")
            return False
        print("  ⚠️  请输入 y 或 n")


def prompt_confirm_execute_after_preview(config=None):
    from multi_exporter import EXPORT_FORMATS
    export_fmt = "excel"
    if config:
        export_fmt = config.get("export_format", "excel")
    fmt_label = EXPORT_FORMATS.get(export_fmt, {}).get("label", export_fmt.upper())

    while True:
        user_input = input(f"\n预览完成！是否导出筛选后的数据到 {fmt_label}？ (Y/n): ").strip().lower()
        if not user_input or user_input in ("y", "yes"):
            return True
        if user_input in ("n", "no"):
            return False
        print("  ⚠️  请输入 y 或 n")


def _to_numeric(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _aggregate_values(values, aggregate_func):
    numeric_values = [v for v in values if _to_numeric(v) is not None]
    nums = [_to_numeric(v) for v in numeric_values]

    if aggregate_func == "sum":
        return sum(nums) if nums else 0
    elif aggregate_func == "count":
        return len(values)
    elif aggregate_func == "count_num":
        return len(nums)
    elif aggregate_func == "average":
        return sum(nums) / len(nums) if nums else 0
    elif aggregate_func == "max":
        return max(nums) if nums else 0
    elif aggregate_func == "min":
        return min(nums) if nums else 0
    elif aggregate_func == "product":
        if not nums:
            return 0
        result = 1
        for n in nums:
            result *= n
        return result
    elif aggregate_func == "stddev":
        if len(nums) < 2:
            return 0
        mean = sum(nums) / len(nums)
        variance = sum((x - mean) ** 2 for x in nums) / (len(nums) - 1)
        return variance ** 0.5
    elif aggregate_func == "stddevp":
        if not nums:
            return 0
        mean = sum(nums) / len(nums)
        variance = sum((x - mean) ** 2 for x in nums) / len(nums)
        return variance ** 0.5
    elif aggregate_func == "var":
        if len(nums) < 2:
            return 0
        mean = sum(nums) / len(nums)
        return sum((x - mean) ** 2 for x in nums) / (len(nums) - 1)
    elif aggregate_func == "varp":
        if not nums:
            return 0
        mean = sum(nums) / len(nums)
        return sum((x - mean) ** 2 for x in nums) / len(nums)
    else:
        return sum(nums) if nums else 0


def _get_field_label(headers, field_key):
    for h in headers:
        if h["key"] == field_key:
            return h["label"]
    return field_key


def _get_row_key(item, row_fields, empty_label):
    key_parts = []
    for field in row_fields:
        val = extract_value(item, field)
        if val is None or val == "":
            val = empty_label
        key_parts.append(str(val))
    return tuple(key_parts)


def _get_col_key(item, col_fields, empty_label):
    if not col_fields:
        return ("",)
    key_parts = []
    for field in col_fields:
        val = extract_value(item, field)
        if val is None or val == "":
            val = empty_label
        key_parts.append(str(val))
    return tuple(key_parts)


def build_pivot_table(data, pivot_config, headers):
    row_fields = pivot_config.get("row_fields", [])
    col_fields = pivot_config.get("column_fields", [])
    value_fields = pivot_config.get("value_fields", [])
    empty_label = pivot_config.get("empty_value_label", "(空白)")

    row_keys = []
    row_keys_order = []
    col_keys = []
    col_keys_order = []

    pivot_data = {}

    for item in data:
        if not isinstance(item, dict):
            continue

        rk = _get_row_key(item, row_fields, empty_label)
        ck = _get_col_key(item, col_fields, empty_label)

        if rk not in row_keys_order:
            row_keys_order.append(rk)
        if ck not in col_keys_order:
            col_keys_order.append(ck)

        if rk not in pivot_data:
            pivot_data[rk] = {}
        if ck not in pivot_data[rk]:
            pivot_data[rk][ck] = {}

        for vf in value_fields:
            field_name = vf["field"]
            agg_func = vf.get("aggregate", "sum")
            vf_key = f"{field_name}:{agg_func}"

            if vf_key not in pivot_data[rk][ck]:
                pivot_data[rk][ck][vf_key] = []

            val = extract_value(item, field_name)
            pivot_data[rk][ck][vf_key].append(val)

    result = {
        "row_fields": row_fields,
        "col_fields": col_fields,
        "value_fields": value_fields,
        "row_keys": row_keys_order,
        "col_keys": col_keys_order,
        "data": pivot_data,
        "empty_label": empty_label,
    }

    return result


def _compute_pivot_value(pivot_result, rk, ck, vf):
    field_name = vf["field"]
    agg_func = vf.get("aggregate", "sum")
    vf_key = f"{field_name}:{agg_func}"

    data = pivot_result["data"]
    if rk in data and ck in data[rk] and vf_key in data[rk][ck]:
        values = data[rk][ck][vf_key]
        return _aggregate_values(values, agg_func)
    return 0


def _compute_row_total(pivot_result, rk, vf):
    total = 0
    agg_func = vf.get("aggregate", "sum")

    if agg_func in ("count", "count_num", "sum"):
        for ck in pivot_result["col_keys"]:
            total += _compute_pivot_value(pivot_result, rk, ck, vf)
    elif agg_func == "average":
        all_values = []
        field_name = vf["field"]
        vf_key = f"{field_name}:{agg_func}"
        data = pivot_result["data"]
        for ck in pivot_result["col_keys"]:
            if rk in data and ck in data[rk] and vf_key in data[rk][ck]:
                all_values.extend(data[rk][ck][vf_key])
        total = _aggregate_values(all_values, agg_func)
    elif agg_func == "max":
        max_val = None
        for ck in pivot_result["col_keys"]:
            v = _compute_pivot_value(pivot_result, rk, ck, vf)
            if max_val is None or v > max_val:
                max_val = v
        total = max_val if max_val is not None else 0
    elif agg_func == "min":
        min_val = None
        for ck in pivot_result["col_keys"]:
            v = _compute_pivot_value(pivot_result, rk, ck, vf)
            if min_val is None or v < min_val:
                min_val = v
        total = min_val if min_val is not None else 0
    else:
        for ck in pivot_result["col_keys"]:
            total += _compute_pivot_value(pivot_result, rk, ck, vf)

    return total


def _compute_col_total(pivot_result, ck, vf):
    total = 0
    agg_func = vf.get("aggregate", "sum")

    if agg_func in ("count", "count_num", "sum"):
        for rk in pivot_result["row_keys"]:
            total += _compute_pivot_value(pivot_result, rk, ck, vf)
    elif agg_func == "average":
        all_values = []
        field_name = vf["field"]
        vf_key = f"{field_name}:{agg_func}"
        data = pivot_result["data"]
        for rk in pivot_result["row_keys"]:
            if rk in data and ck in data[rk] and vf_key in data[rk][ck]:
                all_values.extend(data[rk][ck][vf_key])
        total = _aggregate_values(all_values, agg_func)
    elif agg_func == "max":
        max_val = None
        for rk in pivot_result["row_keys"]:
            v = _compute_pivot_value(pivot_result, rk, ck, vf)
            if max_val is None or v > max_val:
                max_val = v
        total = max_val if max_val is not None else 0
    elif agg_func == "min":
        min_val = None
        for rk in pivot_result["row_keys"]:
            v = _compute_pivot_value(pivot_result, rk, ck, vf)
            if min_val is None or v < min_val:
                min_val = v
        total = min_val if min_val is not None else 0
    else:
        for rk in pivot_result["row_keys"]:
            total += _compute_pivot_value(pivot_result, rk, ck, vf)

    return total


def _compute_grand_total(pivot_result, vf):
    total = 0
    agg_func = vf.get("aggregate", "sum")

    if agg_func in ("count", "count_num", "sum"):
        for rk in pivot_result["row_keys"]:
            total += _compute_row_total(pivot_result, rk, vf)
    elif agg_func == "average":
        all_values = []
        field_name = vf["field"]
        vf_key = f"{field_name}:{agg_func}"
        data = pivot_result["data"]
        for rk in pivot_result["row_keys"]:
            for ck in pivot_result["col_keys"]:
                if rk in data and ck in data[rk] and vf_key in data[rk][ck]:
                    all_values.extend(data[rk][ck][vf_key])
        total = _aggregate_values(all_values, agg_func)
    elif agg_func == "max":
        max_val = None
        for rk in pivot_result["row_keys"]:
            v = _compute_row_total(pivot_result, rk, vf)
            if max_val is None or v > max_val:
                max_val = v
        total = max_val if max_val is not None else 0
    elif agg_func == "min":
        min_val = None
        for rk in pivot_result["row_keys"]:
            v = _compute_row_total(pivot_result, rk, vf)
            if min_val is None or v < min_val:
                min_val = v
        total = min_val if min_val is not None else 0
    else:
        for rk in pivot_result["row_keys"]:
            total += _compute_row_total(pivot_result, rk, vf)

    return total


def _format_value(value, agg_func):
    if isinstance(value, float):
        if value == int(value):
            return int(value)
        return round(value, 2)
    return value


def create_pivot_sheet(wb, pivot_result, pivot_config, headers):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    sheet_name = pivot_config.get("sheet_name", "数据透视表")
    safe_name = _sanitize_sheet_name(sheet_name)
    ws = wb.create_sheet(safe_name)

    row_fields = pivot_result["row_fields"]
    col_fields = pivot_result["col_fields"]
    value_fields = pivot_result["value_fields"]
    row_keys = pivot_result["row_keys"]
    col_keys = pivot_result["col_keys"]
    show_row_totals = pivot_config.get("show_row_totals", True)
    show_col_totals = pivot_config.get("show_column_totals", True)
    grand_total_label = pivot_config.get("grand_total_label", "总计")
    apply_style = pivot_config.get("apply_style", True)

    num_row_fields = len(row_fields)
    num_col_fields = len(col_fields)
    num_value_fields = len(value_fields)

    has_col_fields = len(col_fields) > 0

    if has_col_fields:
        header_row_count = num_col_fields
        if num_value_fields > 1:
            header_row_count += 1
    else:
        header_row_count = 1 if num_value_fields > 1 else 0

    data_start_row = header_row_count + 1

    if has_col_fields:
        total_col_groups = len(col_keys) + (1 if show_col_totals else 0)
        col_header_end_col = num_row_fields + total_col_groups * num_value_fields

        for ci, cf in enumerate(col_fields):
            for c in range(num_row_fields + 1, col_header_end_col + 1):
                cell = ws.cell(row=ci + 1, column=c)
                if c == num_row_fields + 1:
                    cell.value = _get_field_label(headers, cf)
                if apply_style:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        value_header_row = num_col_fields + 1

        col_start_col = num_row_fields + 1
        for cki, ck in enumerate(col_keys):
            group_start = col_start_col + cki * num_value_fields
            group_end = group_start + num_value_fields - 1

            if num_value_fields > 1:
                for c in range(group_start, group_end + 1):
                    cell = ws.cell(row=num_col_fields, column=c)
                    if c == group_start:
                        cell.value = " / ".join(ck)
                    if apply_style:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

            for vfi, vf in enumerate(value_fields):
                cell = ws.cell(row=value_header_row, column=group_start + vfi)
                if num_value_fields == 1:
                    cell.value = " / ".join(ck)
                else:
                    label = vf.get("label") or _get_field_label(headers, vf["field"])
                    agg_label = _get_aggregate_label(vf.get("aggregate", "sum"))
                    cell.value = f"{label}({agg_label})"
                if apply_style:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        if show_col_totals:
            total_start_col = col_start_col + len(col_keys) * num_value_fields
            total_end_col = total_start_col + num_value_fields - 1

            for c in range(total_start_col, total_end_col + 1):
                cell = ws.cell(row=num_col_fields, column=c)
                if c == total_start_col:
                    cell.value = grand_total_label
                if apply_style:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

            for vfi, vf in enumerate(value_fields):
                cell = ws.cell(row=value_header_row, column=total_start_col + vfi)
                if num_value_fields > 1:
                    label = vf.get("label") or _get_field_label(headers, vf["field"])
                    agg_label = _get_aggregate_label(vf.get("aggregate", "sum"))
                    cell.value = f"{label}({agg_label})"
                else:
                    cell.value = grand_total_label
                if apply_style:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    else:
        if num_value_fields > 1:
            for vfi, vf in enumerate(value_fields):
                cell = ws.cell(row=1, column=num_row_fields + 1 + vfi)
                label = vf.get("label") or _get_field_label(headers, vf["field"])
                agg_label = _get_aggregate_label(vf.get("aggregate", "sum"))
                cell.value = f"{label}({agg_label})"
                if apply_style:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    for rfi, rf in enumerate(row_fields):
        cell = ws.cell(row=data_start_row - 1, column=rfi + 1)
        cell.value = _get_field_label(headers, rf)
        if apply_style:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    if has_col_fields:
        for r in range(1, data_start_row):
            for c in range(1, num_row_fields + 1):
                cell = ws.cell(row=r, column=c)
                if apply_style:
                    cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    for rki, rk in enumerate(row_keys):
        row_idx = data_start_row + rki
        for rfi, rv in enumerate(rk):
            cell = ws.cell(row=row_idx, column=rfi + 1)
            cell.value = rv
            if apply_style:
                cell.alignment = Alignment(vertical="center")
                if rki % 2 == 1:
                    cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

        if has_col_fields:
            col_start_col = num_row_fields + 1
            for cki, ck in enumerate(col_keys):
                if num_value_fields > 1:
                    for vfi, vf in enumerate(value_fields):
                        cell = ws.cell(row=row_idx, column=col_start_col + cki * num_value_fields + vfi)
                        val = _compute_pivot_value(pivot_result, rk, ck, vf)
                        cell.value = _format_value(val, vf.get("aggregate", "sum"))
                        if apply_style:
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            if rki % 2 == 1:
                                cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
                else:
                    cell = ws.cell(row=row_idx, column=col_start_col + cki)
                    vf = value_fields[0]
                    val = _compute_pivot_value(pivot_result, rk, ck, vf)
                    cell.value = _format_value(val, vf.get("aggregate", "sum"))
                    if apply_style:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        if rki % 2 == 1:
                            cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

            if show_row_totals:
                if num_value_fields > 1:
                    total_start_col = col_start_col + len(col_keys) * num_value_fields
                    for vfi, vf in enumerate(value_fields):
                        cell = ws.cell(row=row_idx, column=total_start_col + vfi)
                        val = _compute_row_total(pivot_result, rk, vf)
                        cell.value = _format_value(val, vf.get("aggregate", "sum"))
                        if apply_style:
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                else:
                    total_col = col_start_col + len(col_keys)
                    cell = ws.cell(row=row_idx, column=total_col)
                    vf = value_fields[0]
                    val = _compute_row_total(pivot_result, rk, vf)
                    cell.value = _format_value(val, vf.get("aggregate", "sum"))
                    if apply_style:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        else:
            for vfi, vf in enumerate(value_fields):
                cell = ws.cell(row=row_idx, column=num_row_fields + 1 + vfi)
                val = _compute_pivot_value(pivot_result, rk, ("",), vf)
                cell.value = _format_value(val, vf.get("aggregate", "sum"))
                if apply_style:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if rki % 2 == 1:
                        cell.fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

    if show_col_totals and has_col_fields:
        total_row_idx = data_start_row + len(row_keys)
        label_cell = ws.cell(row=total_row_idx, column=1)
        label_cell.value = grand_total_label
        if num_row_fields > 1:
            pass

        col_start_col = num_row_fields + 1
        for cki, ck in enumerate(col_keys):
            if num_value_fields > 1:
                for vfi, vf in enumerate(value_fields):
                    cell = ws.cell(row=total_row_idx, column=col_start_col + cki * num_value_fields + vfi)
                    val = _compute_col_total(pivot_result, ck, vf)
                    cell.value = _format_value(val, vf.get("aggregate", "sum"))
                    if apply_style:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            else:
                cell = ws.cell(row=total_row_idx, column=col_start_col + cki)
                vf = value_fields[0]
                val = _compute_col_total(pivot_result, ck, vf)
                cell.value = _format_value(val, vf.get("aggregate", "sum"))
                if apply_style:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        if show_row_totals:
            if num_value_fields > 1:
                total_start_col = col_start_col + len(col_keys) * num_value_fields
                for vfi, vf in enumerate(value_fields):
                    cell = ws.cell(row=total_row_idx, column=total_start_col + vfi)
                    val = _compute_grand_total(pivot_result, vf)
                    cell.value = _format_value(val, vf.get("aggregate", "sum"))
                    if apply_style:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            else:
                total_col = col_start_col + len(col_keys)
                cell = ws.cell(row=total_row_idx, column=total_col)
                vf = value_fields[0]
                val = _compute_grand_total(pivot_result, vf)
                cell.value = _format_value(val, vf.get("aggregate", "sum"))
                if apply_style:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        if apply_style:
            for c in range(1, num_row_fields + 1):
                cell = ws.cell(row=total_row_idx, column=c)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    if apply_style:
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )
        max_row = ws.max_row
        max_col = ws.max_column
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = thin_border

    for col_idx in range(1, ws.max_column + 1):
        col_letter = chr(64 + col_idx) if col_idx <= 26 else _get_column_letter(col_idx)
        max_width = 12
        for row_idx in range(1, min(ws.max_row + 1, 100)):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                cell_len = len(str(cell_val))
                if cell_len > max_width:
                    max_width = min(cell_len + 4, 40)
        ws.column_dimensions[col_letter].width = max_width

    ws.freeze_panes = ws.cell(row=data_start_row, column=num_row_fields + 1)

    return ws


def _get_aggregate_label(agg_func):
    labels = {
        "sum": "求和",
        "count": "计数",
        "average": "平均值",
        "max": "最大值",
        "min": "最小值",
        "product": "乘积",
        "count_num": "数值计数",
        "stddev": "标准偏差",
        "stddevp": "总体标准偏差",
        "var": "方差",
        "varp": "总体方差",
    }
    return labels.get(agg_func, agg_func)


def add_pivot_table_to_workbook(wb, data, headers, config):
    pivot_config = config.get("pivot_config", {})
    if not pivot_config.get("enabled"):
        return None

    try:
        pivot_result = build_pivot_table(data, pivot_config, headers)
        ws = create_pivot_sheet(wb, pivot_result, pivot_config, headers)
        print(f"  📊 已生成数据透视表: {ws.title}")
        return ws
    except Exception as e:
        print(f"  ⚠️  数据透视表生成失败: {e}")
        import traceback
        traceback.print_exc()
        return None


if __name__ == "__main__":
    main()
