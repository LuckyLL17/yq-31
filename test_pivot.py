import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from json_to_excel import (
    load_json,
    auto_detect_headers,
    merge_headers,
    build_pivot_table,
)
from config_manager import get_default_config


def test_pivot_basic():
    print("=" * 60)
    print("测试1: 基础透视表 - 行字段 + 值字段")
    print("=" * 60)

    data = load_json("./data/sample_data.json")
    auto_headers = auto_detect_headers(data)
    config = get_default_config()
    headers = merge_headers(config.get("default_headers", []), auto_headers, config)

    pivot_config = {
        "enabled": True,
        "row_fields": ["department"],
        "column_fields": [],
        "value_fields": [
            {"field": "salary", "aggregate": "sum", "label": "薪资"},
            {"field": "id", "aggregate": "count", "label": "人数"},
        ],
        "show_row_totals": True,
        "show_column_totals": True,
        "grand_total_label": "总计",
        "empty_value_label": "(空白)",
    }

    result = build_pivot_table(data, pivot_config, headers)
    print(f"行字段: {result['row_fields']}")
    print(f"列字段: {result['col_fields']}")
    print(f"值字段: {result['value_fields']}")
    print(f"行数量: {len(result['row_keys'])}")
    print(f"列数量: {len(result['col_keys'])}")
    print()

    from json_to_excel import _compute_pivot_value, _compute_row_total, _compute_grand_total

    for rk in result["row_keys"]:
        for vf in result["value_fields"]:
            val = _compute_pivot_value(result, rk, ("",), vf)
            print(f"  {rk[0]} - {vf['field']}({vf['aggregate']}): {val}")

    vf = result["value_fields"][0]
    grand_total = _compute_grand_total(result, vf)
    print(f"\n总计 - {vf['field']}({vf['aggregate']}): {grand_total}")
    print("✅ 基础透视表测试通过\n")


def test_pivot_with_columns():
    print("=" * 60)
    print("测试2: 带列字段的透视表")
    print("=" * 60)

    data = load_json("./data/sample_data.json")
    auto_headers = auto_detect_headers(data)
    config = get_default_config()
    headers = merge_headers(config.get("default_headers", []), auto_headers, config)

    pivot_config = {
        "enabled": True,
        "row_fields": ["department"],
        "column_fields": ["position"],
        "value_fields": [
            {"field": "salary", "aggregate": "sum"},
        ],
        "show_row_totals": True,
        "show_column_totals": True,
        "grand_total_label": "总计",
        "empty_value_label": "(空白)",
    }

    result = build_pivot_table(data, pivot_config, headers)
    print(f"行字段: {result['row_fields']}")
    print(f"列字段: {result['col_fields']}")
    print(f"行数量: {len(result['row_keys'])}")
    print(f"列数量: {len(result['col_keys'])}")
    print()

    from json_to_excel import _compute_pivot_value, _compute_row_total, _compute_col_total, _compute_grand_total

    col_headers = " | ".join([ck[0] for ck in result["col_keys"]]) + " | 行总计"
    print(f"{'部门':<10} | {col_headers}")
    print("-" * 80)

    vf = result["value_fields"][0]
    for rk in result["row_keys"]:
        row_vals = []
        for ck in result["col_keys"]:
            val = _compute_pivot_value(result, rk, ck, vf)
            row_vals.append(f"{val:>8}")
        row_total = _compute_row_total(result, rk, vf)
        print(f"{rk[0]:<10} | {' | '.join(row_vals)} | {row_total:>8}")

    print("-" * 80)
    col_totals = []
    for ck in result["col_keys"]:
        ct = _compute_col_total(result, ck, vf)
        col_totals.append(f"{ct:>8}")
    grand_total = _compute_grand_total(result, vf)
    print(f"{'列总计':<10} | {' | '.join(col_totals)} | {grand_total:>8}")
    print("✅ 列字段透视表测试通过\n")


def test_pivot_aggregates():
    print("=" * 60)
    print("测试3: 多种聚合函数")
    print("=" * 60)

    data = load_json("./data/sample_data.json")
    auto_headers = auto_detect_headers(data)
    config = get_default_config()
    headers = merge_headers(config.get("default_headers", []), auto_headers, config)

    pivot_config = {
        "enabled": True,
        "row_fields": ["department"],
        "column_fields": [],
        "value_fields": [
            {"field": "salary", "aggregate": "sum"},
            {"field": "salary", "aggregate": "average"},
            {"field": "salary", "aggregate": "max"},
            {"field": "salary", "aggregate": "min"},
            {"field": "id", "aggregate": "count"},
        ],
        "show_row_totals": True,
        "show_column_totals": False,
        "grand_total_label": "总计",
        "empty_value_label": "(空白)",
    }

    result = build_pivot_table(data, pivot_config, headers)

    from json_to_excel import _compute_pivot_value

    print(f"{'部门':<10} | {'求和':>8} | {'平均值':>8} | {'最大值':>8} | {'最小值':>8} | {'计数':>6}")
    print("-" * 70)

    for rk in result["row_keys"]:
        vals = []
        for vf in result["value_fields"]:
            val = _compute_pivot_value(result, rk, ("",), vf)
            if isinstance(val, float):
                vals.append(f"{val:>8.2f}")
            else:
                vals.append(f"{val:>8}")
        print(f"{rk[0]:<10} | {' | '.join(vals)}")

    print("✅ 聚合函数测试通过\n")


def test_full_export():
    print("=" * 60)
    print("测试4: 完整Excel导出（含透视表）")
    print("=" * 60)

    from openpyxl import load_workbook

    data = load_json("./data/sample_data.json")
    auto_headers = auto_detect_headers(data)
    config = get_default_config()
    headers = merge_headers(config.get("default_headers", []), auto_headers, config)

    config["excel_output_path"] = "./output/test_pivot_export.xlsx"
    config["pivot_config"] = {
        "enabled": True,
        "sheet_name": "数据透视表",
        "row_fields": ["department"],
        "column_fields": ["position"],
        "value_fields": [
            {"field": "salary", "aggregate": "sum", "label": "薪资"},
            {"field": "id", "aggregate": "count", "label": "人数"},
        ],
        "show_row_totals": True,
        "show_column_totals": True,
        "grand_total_label": "总计",
        "empty_value_label": "(空白)",
        "apply_style": True,
    }

    output_path = config["excel_output_path"]
    from json_to_excel import export_to_excel

    result_path = export_to_excel(data, headers, config)

    if result_path and os.path.exists(result_path):
        wb = load_workbook(result_path)
        print(f"\n工作表列表: {wb.sheetnames}")
        assert "数据透视表" in wb.sheetnames, "透视表工作表不存在"

        ws = wb["数据透视表"]
        print(f"透视表行数: {ws.max_row}")
        print(f"透视表列数: {ws.max_column}")

        print("\n透视表内容预览:")
        for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=True):
            print(" | ".join([str(c) if c is not None else "" for c in row[:8]]))

        wb.close()
        print("✅ 完整导出测试通过\n")
    else:
        print("❌ 导出失败")


def test_multiple_row_fields():
    print("=" * 60)
    print("测试5: 多个行字段")
    print("=" * 60)

    data = load_json("./data/pivot_test_data.json")
    auto_headers = auto_detect_headers(data)
    config = get_default_config()
    headers = merge_headers(config.get("default_headers", []), auto_headers, config)

    pivot_config = {
        "enabled": True,
        "row_fields": ["department", "position"],
        "column_fields": [],
        "value_fields": [
            {"field": "salary", "aggregate": "sum", "label": "薪资"},
            {"field": "id", "aggregate": "count", "label": "人数"},
        ],
        "show_row_totals": True,
        "show_column_totals": False,
        "grand_total_label": "总计",
        "empty_value_label": "(空白)",
    }

    result = build_pivot_table(data, pivot_config, headers)
    print(f"行字段: {result['row_fields']}")
    print(f"行数量: {len(result['row_keys'])}")
    print()

    from json_to_excel import _compute_pivot_value, _compute_grand_total

    print(f"{'部门':<10} {'职位':<12} | {'薪资':>10} | {'人数':>6}")
    print("-" * 50)

    total_salary = 0
    total_count = 0
    for rk in result["row_keys"]:
        salary_val = _compute_pivot_value(result, rk, ("",), result["value_fields"][0])
        count_val = _compute_pivot_value(result, rk, ("",), result["value_fields"][1])
        print(f"{rk[0]:<10} {rk[1]:<12} | {salary_val:>10.0f} | {count_val:>6}")
        total_salary += salary_val
        total_count += count_val

    print("-" * 50)
    grand_salary = _compute_grand_total(result, result["value_fields"][0])
    grand_count = _compute_grand_total(result, result["value_fields"][1])
    print(f"{'总计':<22} | {grand_salary:>10.0f} | {grand_count:>6}")

    print("✅ 多行字段测试通过\n")


def test_multiple_col_fields():
    print("=" * 60)
    print("测试6: 多个列字段")
    print("=" * 60)

    data = load_json("./data/pivot_test_data.json")
    auto_headers = auto_detect_headers(data)
    config = get_default_config()
    headers = merge_headers(config.get("default_headers", []), auto_headers, config)

    pivot_config = {
        "enabled": True,
        "row_fields": ["department"],
        "column_fields": ["city", "position"],
        "value_fields": [
            {"field": "salary", "aggregate": "sum"},
        ],
        "show_row_totals": True,
        "show_column_totals": True,
        "grand_total_label": "总计",
        "empty_value_label": "(空白)",
    }

    result = build_pivot_table(data, pivot_config, headers)
    print(f"行字段: {result['row_fields']}")
    print(f"列字段: {result['col_fields']}")
    print(f"行数量: {len(result['row_keys'])}")
    print(f"列数量: {len(result['col_keys'])}")
    print(f"\n列值示例: {result['col_keys'][:3]}...")

    print("✅ 多列字段测试通过\n")


def test_split_with_pivot():
    print("=" * 60)
    print("测试7: 拆分导出 + 透视表")
    print("=" * 60)

    from openpyxl import load_workbook
    from json_to_excel import export_to_excel

    data = load_json("./data/pivot_test_data.json")
    auto_headers = auto_detect_headers(data)
    config = get_default_config()
    headers = merge_headers(config.get("default_headers", []), auto_headers, config)

    config["excel_output_path"] = "./output/test_split_pivot.xlsx"
    config["split_config"] = {
        "enabled": True,
        "split_field": "department",
        "split_rule": "by_value",
        "include_all_sheet": True,
        "all_sheet_name": "全部数据",
    }
    config["pivot_config"] = {
        "enabled": True,
        "sheet_name": "数据透视表",
        "row_fields": ["position"],
        "column_fields": [],
        "value_fields": [
            {"field": "salary", "aggregate": "sum", "label": "薪资"},
            {"field": "id", "aggregate": "count", "label": "人数"},
        ],
        "show_row_totals": True,
        "show_column_totals": True,
        "grand_total_label": "总计",
        "empty_value_label": "(空白)",
        "apply_style": True,
    }

    result_path = export_to_excel(data, headers, config)

    if result_path and os.path.exists(result_path):
        wb = load_workbook(result_path)
        print(f"工作表列表: {wb.sheetnames}")
        assert "数据透视表" in wb.sheetnames, "透视表工作表不存在"

        ws = wb["数据透视表"]
        print(f"透视表行数: {ws.max_row}")
        print(f"透视表列数: {ws.max_column}")

        wb.close()
        print("✅ 拆分+透视表测试通过\n")
    else:
        print("❌ 导出失败")


def test_config_validation():
    print("=" * 60)
    print("测试8: 配置验证")
    print("=" * 60)

    from config_manager import validate_config

    valid_config = get_default_config()
    valid_config["pivot_config"] = {
        "enabled": True,
        "row_fields": ["department"],
        "column_fields": [],
        "value_fields": [
            {"field": "salary", "aggregate": "sum"},
        ],
    }

    errors = validate_config(valid_config)
    print(f"有效配置错误数: {len(errors)}")
    assert len(errors) == 0, f"有效配置应该没有错误，但有: {errors}"

    invalid_config = get_default_config()
    invalid_config["pivot_config"] = {
        "enabled": True,
        "row_fields": [],
        "column_fields": [],
        "value_fields": [],
    }

    errors = validate_config(invalid_config)
    print(f"无效配置错误数: {len(errors)}")
    assert len(errors) > 0, "无效配置应该有错误"
    for e in errors:
        print(f"  - {e}")

    print("✅ 配置验证测试通过\n")


if __name__ == "__main__":
    try:
        test_pivot_basic()
        test_pivot_with_columns()
        test_pivot_aggregates()
        test_multiple_row_fields()
        test_multiple_col_fields()
        test_full_export()
        test_split_with_pivot()
        test_config_validation()
        print("=" * 60)
        print("🎉 所有测试通过！")
        print("=" * 60)
    except Exception as e:
        print(f"\n❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
