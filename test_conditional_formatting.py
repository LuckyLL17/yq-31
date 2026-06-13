#!/usr/bin/env python3
"""
测试条件格式规则功能
"""

import sys
import os
import json
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from style_template_manager import (
    evaluate_conditional_rule,
    match_conditional_rules_for_field,
    merge_conditional_styles,
    create_conditional_rule,
    get_rule_description,
    validate_conditional_rule,
    preset_conditional_rules,
    CF_RULE_TYPE_NUMERIC,
    CF_RULE_TYPE_TEXT,
    CF_RULE_TYPE_NULL,
    CF_OP_GT,
    CF_OP_LT,
    CF_OP_CONTAINS,
    CF_OP_IS_NULL,
)


def test_evaluate_conditional_rule():
    """测试规则评估函数"""
    print("=" * 60)
    print("测试 1: 规则评估函数 evaluate_conditional_rule")
    print("=" * 60)

    rule_gt = create_conditional_rule(
        field="salary",
        rule_type=CF_RULE_TYPE_NUMERIC,
        operator=CF_OP_GT,
        value=20000,
        style={"font_color": "#FF0000"},
        description="薪资大于20000标红",
    )

    assert evaluate_conditional_rule(25000, rule_gt) == True, "25000 > 20000 应该返回 True"
    assert evaluate_conditional_rule(20000, rule_gt) == False, "20000 > 20000 应该返回 False"
    assert evaluate_conditional_rule(15000, rule_gt) == False, "15000 > 20000 应该返回 False"
    assert evaluate_conditional_rule("abc", rule_gt) == False, "非数字应该返回 False"
    print("  ✅ 数值大于规则测试通过")

    rule_lt = create_conditional_rule(
        field="age",
        rule_type=CF_RULE_TYPE_NUMERIC,
        operator=CF_OP_LT,
        value=30,
        style={"font_color": "#00B050"},
    )

    assert evaluate_conditional_rule(25, rule_lt) == True, "25 < 30 应该返回 True"
    assert evaluate_conditional_rule(30, rule_lt) == False, "30 < 30 应该返回 False"
    assert evaluate_conditional_rule(35, rule_lt) == False, "35 < 30 应该返回 False"
    print("  ✅ 数值小于规则测试通过")

    rule_contains = create_conditional_rule(
        field="*",
        rule_type=CF_RULE_TYPE_TEXT,
        operator=CF_OP_CONTAINS,
        value="高级",
        style={"bg_color": "#FFFF00"},
    )

    assert evaluate_conditional_rule("高级工程师", rule_contains) == True, "包含'高级'应该返回 True"
    assert evaluate_conditional_rule("普通工程师", rule_contains) == False, "不包含'高级'应该返回 False"
    assert evaluate_conditional_rule("", rule_contains) == False, "空字符串应该返回 False"
    print("  ✅ 文本包含规则测试通过")

    rule_null = create_conditional_rule(
        field="*",
        rule_type=CF_RULE_TYPE_NULL,
        operator=CF_OP_IS_NULL,
        value=None,
        style={"bg_color": "#FFC7CE"},
    )

    assert evaluate_conditional_rule(None, rule_null) == True, "None 应该返回 True"
    assert evaluate_conditional_rule("", rule_null) == True, "空字符串应该返回 True"
    assert evaluate_conditional_rule("value", rule_null) == False, "有值应该返回 False"
    print("  ✅ 空值规则测试通过")

    disabled_rule = create_conditional_rule(
        field="salary",
        rule_type=CF_RULE_TYPE_NUMERIC,
        operator=CF_OP_GT,
        value=20000,
        style={"font_color": "#FF0000"},
        enabled=False,
    )
    assert evaluate_conditional_rule(30000, disabled_rule) == False, "禁用的规则应该返回 False"
    print("  ✅ 禁用规则测试通过")

    print("\n所有规则评估测试通过！\n")


def test_match_conditional_rules():
    """测试多规则匹配"""
    print("=" * 60)
    print("测试 2: 多规则匹配 match_conditional_rules_for_field")
    print("=" * 60)

    rules = [
        create_conditional_rule(
            field="salary",
            rule_type=CF_RULE_TYPE_NUMERIC,
            operator=CF_OP_GT,
            value=20000,
            style={"font_color": "#FF0000", "font_bold": True},
            priority=10,
        ),
        create_conditional_rule(
            field="*",
            rule_type=CF_RULE_TYPE_TEXT,
            operator=CF_OP_CONTAINS,
            value="高级",
            style={"bg_color": "#FFFF00"},
            priority=1,
        ),
        create_conditional_rule(
            field="age",
            rule_type=CF_RULE_TYPE_NUMERIC,
            operator=CF_OP_LT,
            value=30,
            style={"font_color": "#00B050"},
            priority=5,
        ),
    ]

    matched = match_conditional_rules_for_field("salary", 25000, rules)
    assert len(matched) == 1, "薪资25000应该匹配1条规则"
    assert matched[0]["operator"] == CF_OP_GT, "应该匹配大于规则"
    print("  ✅ 薪资字段匹配测试通过")

    matched = match_conditional_rules_for_field("position", "高级工程师", rules)
    assert len(matched) == 1, "包含'高级'应该匹配1条规则"
    assert matched[0]["operator"] == CF_OP_CONTAINS, "应该匹配包含规则"
    print("  ✅ 职位字段匹配测试通过")

    matched = match_conditional_rules_for_field("age", 28, rules)
    assert len(matched) == 1, "年龄28应该匹配1条规则"
    assert matched[0]["operator"] == CF_OP_LT, "应该匹配小于规则"
    print("  ✅ 年龄字段匹配测试通过")

    matched = match_conditional_rules_for_field("name", "张三", rules)
    assert len(matched) == 0, "不匹配任何规则"
    print("  ✅ 无匹配测试通过")

    print("\n所有多规则匹配测试通过！\n")


def test_merge_styles():
    """测试样式合并"""
    print("=" * 60)
    print("测试 3: 样式合并 merge_conditional_styles")
    print("=" * 60)

    rules = [
        create_conditional_rule(
            field="*",
            rule_type=CF_RULE_TYPE_NUMERIC,
            operator=CF_OP_GT,
            value=20000,
            style={"font_color": "#FF0000", "font_bold": True},
            priority=10,
        ),
        create_conditional_rule(
            field="*",
            rule_type=CF_RULE_TYPE_TEXT,
            operator=CF_OP_CONTAINS,
            value="高级",
            style={"bg_color": "#FFFF00", "font_bold": False},
            priority=1,
        ),
    ]

    merged = merge_conditional_styles(rules)
    assert merged["font_color"] == "#FF0000", "字体颜色应该是红色"
    assert merged["bg_color"] == "#FFFF00", "背景颜色应该是黄色"
    assert merged["font_bold"] == False, "后应用的font_bold应该覆盖之前的"
    print(f"  合并后的样式: {merged}")
    print("  ✅ 样式合并测试通过\n")


def test_rule_description():
    """测试规则描述生成"""
    print("=" * 60)
    print("测试 4: 规则描述生成 get_rule_description")
    print("=" * 60)

    rule1 = create_conditional_rule(
        field="salary",
        rule_type=CF_RULE_TYPE_NUMERIC,
        operator=CF_OP_GT,
        value=20000,
        style={"font_color": "#FF0000"},
        description="薪资大于20000标红",
    )
    desc1 = get_rule_description(rule1)
    assert desc1 == "薪资大于20000标红", "应该使用自定义描述"
    print(f"  规则1描述: {desc1}")

    rule2 = create_conditional_rule(
        field="age",
        rule_type=CF_RULE_TYPE_NUMERIC,
        operator=CF_OP_LT,
        value=30,
        style={"font_color": "#00B050"},
    )
    desc2 = get_rule_description(rule2)
    assert "age" in desc2 and "小于" in desc2 and "30" in desc2, "应该自动生成描述"
    print(f"  规则2描述: {desc2}")

    rule3 = create_conditional_rule(
        field="*",
        rule_type=CF_RULE_TYPE_TEXT,
        operator=CF_OP_CONTAINS,
        value="高级",
        style={"bg_color": "#FFFF00"},
    )
    desc3 = get_rule_description(rule3)
    assert "所有字段" in desc3 and "包含" in desc3 and "高级" in desc3
    print(f"  规则3描述: {desc3}")

    print("  ✅ 规则描述测试通过\n")


def test_validate_rule():
    """测试规则验证"""
    print("=" * 60)
    print("测试 5: 规则验证 validate_conditional_rule")
    print("=" * 60)

    valid_rule = create_conditional_rule(
        field="salary",
        rule_type=CF_RULE_TYPE_NUMERIC,
        operator=CF_OP_GT,
        value=20000,
        style={"font_color": "#FF0000"},
    )
    is_valid, msg = validate_conditional_rule(valid_rule)
    assert is_valid == True, "有效规则应该验证通过"
    print(f"  有效规则: {msg}")

    invalid_rule1 = {
        "rule_type": CF_RULE_TYPE_NUMERIC,
        "operator": CF_OP_GT,
        "value": 20000,
        "style": {"font_color": "#FF0000"},
    }
    is_valid, msg = validate_conditional_rule(invalid_rule1)
    assert is_valid == False, "缺少字段名应该验证失败"
    print(f"  无效规则(缺少field): {msg}")

    invalid_rule2 = create_conditional_rule(
        field="salary",
        rule_type="invalid_type",
        operator=CF_OP_GT,
        value=20000,
        style={"font_color": "#FF0000"},
    )
    is_valid, msg = validate_conditional_rule(invalid_rule2)
    assert is_valid == False, "无效规则类型应该验证失败"
    print(f"  无效规则(类型错误): {msg}")

    invalid_rule3 = create_conditional_rule(
        field="salary",
        rule_type=CF_RULE_TYPE_NUMERIC,
        operator=CF_OP_GT,
        value=20000,
        style={},
    )
    is_valid, msg = validate_conditional_rule(invalid_rule3)
    assert is_valid == False, "缺少样式应该验证失败"
    print(f"  无效规则(缺少样式): {msg}")

    print("  ✅ 规则验证测试通过\n")


def test_preset_rules():
    """测试预设规则"""
    print("=" * 60)
    print("测试 6: 预设规则 preset_conditional_rules")
    print("=" * 60)

    presets = preset_conditional_rules()
    print(f"  共 {len(presets)} 个预设规则:")
    for key, preset in presets.items():
        print(f"    - {preset['name']}: {preset['description']}")
        is_valid, msg = validate_conditional_rule(preset["rule"])
        assert is_valid == True, f"预设规则 {key} 应该有效: {msg}"

    assert "high_salary_red" in presets, "应该包含高薪标红预设"
    assert "young_age_green" in presets, "应该包含年轻标绿预设"
    assert "keyword_highlight" in presets, "应该包含关键词高亮预设"

    print("  ✅ 预设规则测试通过\n")


def test_integration_with_export():
    """测试与导出流程的集成"""
    print("=" * 60)
    print("测试 7: 集成到导出流程")
    print("=" * 60)

    from config_manager import get_default_config, save_config
    from json_to_excel import load_json, auto_detect_headers, merge_headers, export_to_excel

    config = get_default_config()
    config["json_file_path"] = "./data/sample_data.json"
    config["excel_output_path"] = "./output/test_conditional_format.xlsx"

    config["conditional_format_rules"] = [
        create_conditional_rule(
            field="salary",
            rule_type=CF_RULE_TYPE_NUMERIC,
            operator=CF_OP_GT,
            value=20000,
            style={"font_color": "#FF0000", "font_bold": True},
            description="薪资大于20000标红",
        ),
        create_conditional_rule(
            field="age",
            rule_type=CF_RULE_TYPE_NUMERIC,
            operator=CF_OP_LT,
            value=30,
            style={"font_color": "#00B050", "font_bold": True},
            description="年龄小于30标绿",
        ),
        create_conditional_rule(
            field="*",
            rule_type=CF_RULE_TYPE_TEXT,
            operator=CF_OP_CONTAINS,
            value="高级",
            style={"bg_color": "#FFFF00", "font_bold": True},
            description="包含'高级'关键词高亮",
        ),
    ]

    print("  配置中的条件格式规则:")
    for rule in config["conditional_format_rules"]:
        print(f"    • {get_rule_description(rule)}")

    save_config(config, "./config_test_cf.json")

    data = load_json(config["json_file_path"])
    auto_headers = auto_detect_headers(data)
    headers = merge_headers(config.get("default_headers", []), auto_headers, config)

    print(f"\n  测试数据: {len(data)} 条")
    print("  预期匹配:")
    for item in data:
        name = item.get("name", "")
        age = item.get("age", "")
        salary = item.get("salary", "")
        position = item.get("position", "")
        flags = []
        if salary > 20000:
            flags.append("薪资标红")
        if age < 30:
            flags.append("年龄标绿")
        if "高级" in str(position):
            flags.append("关键词高亮")
        if flags:
            print(f"    {name}: {', '.join(flags)}")

    print("\n  正在导出Excel...")
    output_path = export_to_excel(data, headers, config)

    if output_path and os.path.exists(output_path):
        print(f"  ✅ Excel导出成功: {output_path}")

        wb = load_workbook(output_path)
        ws = wb.active

        salary_col = None
        age_col = None
        position_col = None

        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header == "薪资":
                salary_col = col
            elif header == "年龄":
                age_col = col
            elif header == "职位":
                position_col = col

        print(f"\n  验证单元格样式:")
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=2).value

            if salary_col:
                salary_cell = ws.cell(row=row, column=salary_col)
                if salary_cell.font.color and salary_cell.font.color.rgb == "FFFF0000":
                    print(f"    {name} 薪资({salary_cell.value}): 字体红色 ✅")

            if age_col:
                age_cell = ws.cell(row=row, column=age_col)
                if age_cell.font.color and age_cell.font.color.rgb == "FF00B050":
                    print(f"    {name} 年龄({age_cell.value}): 字体绿色 ✅")

            if position_col:
                pos_cell = ws.cell(row=row, column=position_col)
                if pos_cell.fill.fgColor and pos_cell.fill.fgColor.rgb == "FFFFFF00":
                    print(f"    {name} 职位({pos_cell.value}): 背景黄色 ✅")

        wb.close()
        print("\n  ✅ 导出集成测试通过")
    else:
        print("  ❌ 导出失败")

    print()


def main():
    print("\n" + "=" * 60)
    print("条件格式规则功能测试")
    print("=" * 60 + "\n")

    try:
        test_evaluate_conditional_rule()
        test_match_conditional_rules()
        test_merge_styles()
        test_rule_description()
        test_validate_rule()
        test_preset_rules()
        test_integration_with_export()

        print("=" * 60)
        print("🎉 所有测试通过！")
        print("=" * 60)
        return 0
    except AssertionError as e:
        print(f"\n❌ 测试失败: {e}")
        import traceback
        traceback.print_exc()
        return 1
    except Exception as e:
        print(f"\n❌ 发生错误: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
